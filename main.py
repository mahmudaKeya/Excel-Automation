import openpyxl

from openpyxl import Workbook, load_workbook


wb = load_workbook('G:/Topup_automation/test.xlsx')
ws = wb.active

row = ws.max_row
column = ws.max_column

# row = 21


ws['I1'].value = "Amount"
ws['J1'].value = "Notification"
ws['K1'].value = "TopUp"
ws['L1'].value = "Status"


print(row, column)



for i in range(2, row + 1):
    cell_obj = ws.cell(row = i, column = 7)
    # print(cell_obj.value)

    b = []

    date = []

    put = cell_obj.value

    if put != -1:
        get = put.find("tk")
        print(get)
        tarikh = put.find("th ")

        b.insert(0, put[get - 2])
        b.insert(1, put[get - 1])

        date.insert(0, put[tarikh - 2])
        date.insert(1, put[tarikh - 1])


        mynewlist = [s for s in b if s.isdigit()]
        print(type(mynewlist))
        # mynewlist_date = [s for s in date if s.isdigit()]
        # print("Date", mynewlist_date)

        if len(mynewlist) != 0:
            res = int("".join(map(str, mynewlist)))
            # res_date = int("".join(map(str, mynewlist_date)))
            print("My new value", res)
            # cell_obj.value = res
            ws.cell(row = i, column = 9).value = res
            # ws.cell(row = i, column = 14).value = res_date




        # res = int("".join(map(str, mynewlist)))
        # print("My new List", mynewlist)

        # mynewlist = [s for s in b if s.isdigit()]



    # if len(mynewlist) == 0:
    #     print("list is empty")
    # else:
    #     print("list is not empty")


    # res = int("".join(map(str, mynewlist)))

    print(put)
    print(get)
    print(b)
    # print("My new List", mynewlist)


    # print(type(cell_obj.value))

wb.save('G:/Topup_automation/test.xlsx')
